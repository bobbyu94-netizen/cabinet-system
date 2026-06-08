# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from math import gcd

wb = Workbook()

# ── Colors ────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", start_color="D9D9D9")  # column headers
ALT_FILL = PatternFill("solid", start_color="F2F2F2")  # alternating row
WHT_FILL = PatternFill("solid", start_color="FFFFFF")  # white row
PHS_FILL = PatternFill("solid", start_color="BFBFBF")  # phase headers
SEC_FILL = PatternFill("solid", start_color="E8EEF4")  # section sub-headers
REM_FILL = PatternFill("solid", start_color="EBF1DE")  # remnant row
SCP_FILL = PatternFill("solid", start_color="FCE4D6")  # scrap row
NO_FILL  = PatternFill("solid", start_color="FFFFFF")

FONT = "Arial"

def f(size=13, bold=False, italic=False):
    return Font(name=FONT, size=size, bold=bold, italic=italic)

def to_frac(inches):
    """Convert decimal inches to fractional string, e.g. 34.5 -> '34-1/2\"'"""
    whole = int(inches)
    frac = round((inches - whole) * 16)
    if frac == 0:
        return f'{whole}"'
    if frac == 16:
        return f'{whole + 1}"'
    g = gcd(frac, 16)
    num, den = frac // g, 16 // g
    return f'{whole}-{num}/{den}"' if whole > 0 else f'{num}/{den}"'

def setup_page(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage   = True

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_cell(ws, row, col, value, fnt=None, fill=None, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    if fnt:  c.font = fnt
    if fill: c.fill = fill
    c.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    return c

def sheet_header(ws, row, title, note):
    write_cell(ws, row,   1, title, fnt=f(14, bold=True))
    write_cell(ws, row+1, 1, note,  fnt=f(13, italic=True))
    # row+2 is spacer

def col_hdr_row(ws, row, labels, num_cols):
    for i, lbl in enumerate(labels, 1):
        write_cell(ws, row, i, lbl, fnt=f(13, bold=True), fill=HDR_FILL)
    for i in range(len(labels)+1, num_cols+1):
        ws.cell(row=row, column=i).fill = HDR_FILL

def data_row(ws, row, values, idx, num_cols, wrap_last=False):
    fill = ALT_FILL if idx % 2 == 0 else WHT_FILL
    for i, v in enumerate(values, 1):
        wrap = wrap_last and (i == len(values))
        write_cell(ws, row, i, v, fnt=f(13), fill=fill, wrap=wrap)
    for i in range(len(values)+1, num_cols+1):
        ws.cell(row=row, column=i).fill = fill

# ── Cabinet dimensions (base cabinet, 24" wide) ───────────────────────────────
# Back: full cabinet height
# Face frame stile: back height - toe kick + 0.75"
# Stile = 34.5 - 4 + 0.75 = 31.25"
# Total opening = 31.25 - (3 rails x 1.5) = 26.75"
# Drawer opening = 5" (6" face - 1" overlay)
# Door opening = 21.75"; door height = 22.75"
# Door width each = (21/2) + 0.5 - 0.0625 = 10.9375"
# Door rail = 10.9375 - 6 + 0.75 = 5.6875"
# MDF panel = 5.6875" x (22.75 - 5.25) = 5.6875" x 17.5"
# Drawer box: width=20", depth=21", height=5/8*5=3.125"
# Drawer bottom: 19.5" x 20.5" (minus 0.5" each dim for dado)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
setup_page(ws1)
set_widths(ws1, [32, 26, 7, 12, 12, 16, 14, 44])
N = 8

sheet_header(ws1, 1, '24" Base Cabinet — Cut List Summary',
             "Prices are estimates — verify before final quote.")
col_hdr_row(ws1, 4, ["Part Name","Material","Qty","Unit","Unit Cost","Extended Cost","Vendor","Notes"], N)

# Poplar: 5.64 LF needed; 1 x 10ft board (10 LF) @ $8.57/LF = $85.70
items = [
    ("3/4\" Birch Plywood",           "3/4\" Birch Plywood",  1, "sheet",     89.00,  89.00, "Home Depot", "Box parts — back, sides, nailers, bottom, shelf"),
    ("1/2\" Birch Plywood",           "1/2\" Birch Plywood",  1, "sheet",     48.00,  48.00, "Home Depot", "Drawer box sides, front, back"),
    ("1/4\" Birch Plywood",           "1/4\" Birch Plywood",  1, "sheet",     35.00,  35.00, "Home Depot", "Drawer bottom"),
    ("1/4\" MDF",                     "1/4\" MDF",            1, "sheet",     22.00,  22.00, "Home Depot", "Door panels"),
    ("1x10 Poplar",                   "Poplar solid lumber",  1, "10ft board", 85.70,  85.70, "Home Depot", "Face frame + door frames (5.64 LF needed @ $8.57/LF)"),
    ("Concealed Soft-Close Hinge",    "Hardware",             4, "each",       4.50,  18.00, "Home Depot", "2 per door"),
    ("Hinge Mounting Plate",          "Hardware",             4, "each",       2.00,   8.00, "Home Depot", "1 per hinge"),
    ("Undermount Drawer Slide 22\"",  "Hardware",             1, "pair",      25.00,  25.00, "Home Depot", "Soft-close undermount"),
    ("Shelf Pin 5mm",                 "Hardware",             4, "each",       0.25,   1.00, "Home Depot", "4 per shelf location"),
    ("Cabinet Pull",                  "Hardware",             2, "each",       8.00,  16.00, "Home Depot", "Door pulls"),
    ("Cabinet Pull",                  "Hardware",             1, "each",       8.00,   8.00, "Home Depot", "Drawer pull"),
    ("Cabinet Leveler",               "Hardware",             4, "each",       1.50,   6.00, "Home Depot", ""),
]
for i, row in enumerate(items):
    r = 5 + i
    data_row(ws1, r, row, i, N, wrap_last=True)
    for col in (5, 6):
        ws1.cell(r, col).number_format = '"$"#,##0.00'

mat_last = 5 + len(items) - 1

cons_r = mat_last + 1
data_row(ws1, cons_r,
    ("Consumables (13%)", "Supplies", 1, "allowance", "", "", "",
     "Glue, screws, sandpaper, tape, finish supplies"),
    len(items), N)
cc = ws1.cell(cons_r, 6, f"=SUM(F5:F{mat_last})*0.13")
cc.font = f(13)
cc.fill = ALT_FILL if len(items) % 2 == 0 else WHT_FILL
cc.number_format = '"$"#,##0.00'

tot_r = cons_r + 1
for col in range(1, N+1):
    ws1.cell(tot_r, col).fill = HDR_FILL
write_cell(ws1, tot_r, 1, "Total Material Cost", fnt=f(13, bold=True), fill=HDR_FILL)
tc = ws1.cell(tot_r, 6, f"=SUM(F5:F{cons_r})")
tc.font = f(13, bold=True)
tc.fill = HDR_FILL
tc.number_format = '"$"#,##0.00'
write_cell(ws1, tot_r, 8, "Low confidence — verify all prices", fnt=f(13, bold=True), fill=HDR_FILL, wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — 3/4 Birch Parts
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("3-4 Birch Parts")
setup_page(ws2)
set_widths(ws2, [26, 13, 14, 7, 20, 42])
N2 = 6

sheet_header(ws2, 1, '3/4" Birch Plywood — Cabinet Box Parts',
             "All dimensions in inches. 1 sheet required (with 15% waste factor).")
col_hdr_row(ws2, 4, ["Part Name","Width (in)","Length (in)","Qty","Material","Notes"], N2)

birch = [
    ("Back",                 22,    34.5,  1, '3/4" Birch', 'Full cabinet height'),
    ("Left Side",            23.25, 34.5,  1, '3/4" Birch', 'Grain vertical'),
    ("Right Side",           23.25, 34.5,  1, '3/4" Birch', 'Grain vertical'),
    ("Top Nailer — Front",   4,     22,    1, '3/4" Birch', ''),
    ("Top Nailer — Back",    4,     22,    1, '3/4" Birch', ''),
    ("Bottom",               22,    22.5,  1, '3/4" Birch', ''),
    ("Shelf",                22,    21.25, 1, '3/4" Birch', 'Adjustable'),
]
for i, row in enumerate(birch):
    data_row(ws2, 5+i, row, i, N2, wrap_last=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — 3/4 Birch Sheet Plan
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3-4 Birch Sheet Plan")
setup_page(ws3)
set_widths(ws3, [10, 58, 14, 12, 42])
N3 = 5

sheet_header(ws3, 1, '3/4" Birch — Sheet Cutting Plan',
             'Sheet size: 48" x 96". All parts nest on 1 sheet.')
col_hdr_row(ws3, 4, ["Sheet #","Parts Nested","Sheet Used %","Waste %","Notes"], N3)

data_row(ws3, 5,
    ["1", "Back, Left Side, Right Side, Top Nailer Front, Top Nailer Back, Bottom, Shelf",
     "76%", "24%",
     "Sides and back are largest parts — nest first. Nailers and shelf fill remaining space."],
    0, N3, wrap_last=True)
ws3.row_dimensions[5].height = 48

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — 1/2 Birch Parts
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("1-2 Birch Parts")
setup_page(ws4)
set_widths(ws4, [26, 13, 14, 7, 20, 42])
N4 = 6

sheet_header(ws4, 1, '1/2" Birch Plywood — Drawer Box Parts',
             'All dimensions in inches. Drawer box: 20"W x 3.125"H x 21"D.')
col_hdr_row(ws4, 4, ["Part Name","Width (in)","Length (in)","Qty","Material","Notes"], N4)

drawer_box = [
    ("Drawer Side — Left",  3.125, 21, 1, '1/2" Birch', 'Grain horizontal'),
    ("Drawer Side — Right", 3.125, 21, 1, '1/2" Birch', 'Grain horizontal'),
    ("Drawer Front",        3.125, 20, 1, '1/2" Birch', ''),
    ("Drawer Back",         3.125, 20, 1, '1/2" Birch', ''),
]
for i, row in enumerate(drawer_box):
    data_row(ws4, 5+i, row, i, N4, wrap_last=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — 1/4 Birch Parts
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("1-4 Birch Parts")
setup_page(ws5)
set_widths(ws5, [26, 13, 14, 7, 20, 42])
N5 = 6

sheet_header(ws5, 1, '1/4" Birch Plywood — Drawer Bottom',
             'Dadoed into drawer box sides. Dimensions account for 0.5" dado depth per axis.')
col_hdr_row(ws5, 4, ["Part Name","Width (in)","Length (in)","Qty","Material","Notes"], N5)

data_row(ws5, 5,
    ("Drawer Bottom", 19.5, 20.5, 1, '1/4" Birch',
     'Box width (20) - 0.5" dado; box depth (21) - 0.5" dado'),
    0, N5, wrap_last=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — 1/4 MDF Door Panels
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("1-4 MDF Door Panels")
setup_page(ws6)
set_widths(ws6, [30, 13, 14, 7, 52])
N6 = 5

# Door opening: 21"W x 21.75"H; each door 10.9375"W x 22.75"H
# Panel: door - 5.25" each axis => 5.6875"W x 17.5"H
sheet_header(ws6, 1, '1/4" MDF — Door Panels',
             '1 sheet required. Door opening: 21"W x 21.75"H. '
             '2 doors at 10-15/16"W x 22-3/4"H. '
             'Overlay: 1/2" hinge side, 1/8" total center gap.')
col_hdr_row(ws6, 4, ["Part Name","Width (in)","Length (in)","Qty","Notes"], N6)

mdf = [
    ("Door Panel — Left Door",  5.6875, 17.5, 1, 'Panel = door width - 5.25", door height - 5.25"'),
    ("Door Panel — Right Door", 5.6875, 17.5, 1, 'Panel = door width - 5.25", door height - 5.25"'),
]
for i, row in enumerate(mdf):
    data_row(ws6, 5+i, row, i, N6, wrap_last=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Poplar Parts
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Poplar Parts")
setup_page(ws7)
set_widths(ws7, [36, 13, 14, 7, 52])
N7 = 5

# Poplar LF: 1.5" group: (2x31.25 + 3x21)/6 = 1.74 LF
#            3.0" group: (4x22.75 + 4x5.6875)/3 = 3.16 LF
#            Total: 4.90 LF x 1.15 waste = 5.64 LF -> 1 x 10ft board
sheet_header(ws7, 1, 'Poplar — Face Frame & Door Frame Parts',
             '1x10 stock (9.0" planning width). 1 x 10ft board required (5.64 LF needed + 15% waste).')
col_hdr_row(ws7, 4, ["Part Name","Width (in)","Length (in)","Qty","Notes"], N7)

poplar_parts = [
    ("Face Frame Stile — Left",         1.5,    31.25,  1, 'Back height (34.5) - toe kick (4) + 0.75 = 31.25"'),
    ("Face Frame Stile — Right",        1.5,    31.25,  1, 'Back height (34.5) - toe kick (4) + 0.75 = 31.25"'),
    ("Face Frame Rail — Top",           1.5,    21.0,   1, ''),
    ("Face Frame Rail — Drawer Divider",1.5,    21.0,   1, 'Separates drawer opening from door opening'),
    ("Face Frame Rail — Bottom",        1.5,    21.0,   1, 'Sits at toe kick level'),
    ("Door Stile — Left Door, Hinge",   3.0,    22.75,  1, ''),
    ("Door Stile — Left Door, Center",  3.0,    22.75,  1, ''),
    ("Door Stile — Right Door, Center", 3.0,    22.75,  1, ''),
    ("Door Stile — Right Door, Hinge",  3.0,    22.75,  1, ''),
    ("Door Rail — Left Door, Top",      3.0,    5.6875, 1, 'Door width - 5.25"'),
    ("Door Rail — Left Door, Bottom",   3.0,    5.6875, 1, 'Door width - 5.25"'),
    ("Door Rail — Right Door, Top",     3.0,    5.6875, 1, 'Door width - 5.25"'),
    ("Door Rail — Right Door, Bottom",  3.0,    5.6875, 1, 'Door width - 5.25"'),
]
for i, row in enumerate(poplar_parts):
    data_row(ws7, 5+i, row, i, N7, wrap_last=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Poplar Cut Plan
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Poplar Cut Plan")
setup_page(ws8)
set_widths(ws8, [36, 24, 44])
N8 = 3

# Board: 1x10 x 10ft (120")
# Sections: A=31.25, B=22.75, C=22.75, D=21, E=5.6875
# Total used: 103.9375" | Remnant: ~16-1/16"
sheet_header(ws8, 1, 'Poplar — Board Breakdown & Cut Plan',
             'Stock: 1x 1x10 x 10ft  |  Planning width: 9.0"  |  All parts from 1 board')

r = 4

def phase_hdr(ws, row, text):
    for col in range(1, N8+1):
        write_cell(ws, row, col, "", fnt=f(13, bold=True), fill=PHS_FILL)
    write_cell(ws, row, 1, text, fnt=f(13, bold=True), fill=PHS_FILL)
    ws.row_dimensions[row].height = 18

def sec_hdr(ws, row, text):
    for col in range(1, N8+1):
        write_cell(ws, row, col, "", fnt=f(13, bold=True), fill=SEC_FILL)
    write_cell(ws, row, 1, text, fnt=f(13, bold=True), fill=SEC_FILL)
    ws.row_dimensions[row].height = 18

def ph_col_hdr(ws, row, labels):
    col_hdr_row(ws, row, labels, N8)

def spacer(ws, row):
    for col in range(1, N8+1):
        ws.cell(row=row, column=col).fill = NO_FILL

# ── PHASE 1 ───────────────────────────────────────────────────────────────────
phase_hdr(ws8, r, "PHASE 1 — MAIN BOARD CROSSCUTS"); r += 1
ph_col_hdr(ws8, r, ["Section", "Crosscut Length", "Purpose"]); r += 1

p1_sections = [
    ("A", 31.25,   "Face frame stiles"),
    ("B", 22.75,   "Door stiles 1 & 2"),
    ("C", 22.75,   "Door stiles 3 & 4"),
    ("D", 21.0,    "Face frame rails (all 3)"),
    ("E", 5.6875,  "Door rails 1 & 2"),
]
for i, (sec, length, purpose) in enumerate(p1_sections):
    data_row(ws8, r, [sec, to_frac(length), purpose], i, N8)
    r += 1

# Remnant row
for col in range(1, N8+1):
    ws8.cell(r, col).fill = REM_FILL
write_cell(ws8, r, 1, "Remnant", fnt=f(13), fill=REM_FILL)
write_cell(ws8, r, 2, '~16-1/16"', fnt=f(13), fill=REM_FILL)
write_cell(ws8, r, 3, "Offcut — save for shop stock", fnt=f(13), fill=REM_FILL)
r += 1
spacer(ws8, r); r += 1

# ── PHASE 2 ───────────────────────────────────────────────────────────────────
phase_hdr(ws8, r, "PHASE 2 — RIP ALL SECTIONS"); r += 1

# Section A — FF Stiles (31.25")
sec_hdr(ws8, r, f'Section A — 9.0" x {to_frac(31.25)}'); r += 1
ph_col_hdr(ws8, r, ["Rip Width", "Part Produced", "Remaining Width"]); r += 1
data_row(ws8, r, ['1.5"', "✓ FF Stile 1",  '7.375"'],                        0, N8); r += 1
data_row(ws8, r, ['1.5"', "✓ FF Stile 2",  'Offcut A: 5.875" x 31-1/4"'],   1, N8); r += 1
spacer(ws8, r); r += 1

# Section B — Door Stiles 1 & 2 (22.75")
sec_hdr(ws8, r, f'Section B — 9.0" x {to_frac(22.75)}'); r += 1
ph_col_hdr(ws8, r, ["Rip Width", "Part Produced", "Remaining Width"]); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Stile 1", '5.875"'],                       0, N8); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Stile 2", 'Offcut B: 2.875" x 22-3/4"'],  1, N8); r += 1
spacer(ws8, r); r += 1

# Section C — Door Stiles 3 & 4 (22.75")
sec_hdr(ws8, r, f'Section C — 9.0" x {to_frac(22.75)}'); r += 1
ph_col_hdr(ws8, r, ["Rip Width", "Part Produced", "Remaining Width"]); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Stile 3", '5.875"'],                       0, N8); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Stile 4", 'Offcut C: 2.875" x 22-3/4"'],  1, N8); r += 1
spacer(ws8, r); r += 1

# Section D — FF Rails x3 (21")
sec_hdr(ws8, r, f'Section D — 9.0" x {to_frac(21.0)}'); r += 1
ph_col_hdr(ws8, r, ["Rip Width", "Part Produced", "Remaining Width"]); r += 1
data_row(ws8, r, ['1.5"', "✓ FF Rail 1 (Top)",           '7.375"'],          0, N8); r += 1
data_row(ws8, r, ['1.5"', "✓ FF Rail 2 (Drawer Divider)", '5.75"'],          1, N8); r += 1
data_row(ws8, r, ['1.5"', "✓ FF Rail 3 (Bottom)",         'Offcut D: 4.125" x 21"'], 0, N8); r += 1
spacer(ws8, r); r += 1

# Section E — Door Rails 1 & 2 (5.6875")
sec_hdr(ws8, r, f'Section E — 9.0" x {to_frac(5.6875)}'); r += 1
ph_col_hdr(ws8, r, ["Rip Width", "Part Produced", "Remaining Width"]); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Rail 1", '5.875"'],                        0, N8); r += 1
data_row(ws8, r, ['3.0"', "✓ Door Rail 2", 'Offcut E: 2.875" x 5-11/16"'],  1, N8); r += 1
spacer(ws8, r); r += 1

# ── PHASE 3 ───────────────────────────────────────────────────────────────────
phase_hdr(ws8, r, "PHASE 3 — OFFCUT CROSSCUTS & RIPS"); r += 1

# Offcut A (5.875" x 31.25") -> Door Rails 3 & 4
sec_hdr(ws8, r, 'Offcut A (5.875" x 31-1/4") — Door Rails 3 & 4'); r += 1
ph_col_hdr(ws8, r, ["Action", "Part Produced", "Result"]); r += 1
data_row(ws8, r, ['Rip 3.0"', "Door Rail Strip", 'Offcut A2: 2.75" x 31-1/4"'], 0, N8); r += 1
spacer(ws8, r); r += 1

sec_hdr(ws8, r, 'Door Rail Strip (3.0" x 31-1/4") — Door Rails 3 & 4'); r += 1
ph_col_hdr(ws8, r, ["Action", "Part Produced", "Remaining"]); r += 1
data_row(ws8, r, ['Crosscut 5-11/16"', "✓ Door Rail 3", '25-7/16"'],              0, N8); r += 1
data_row(ws8, r, ['Crosscut 5-11/16"', "✓ Door Rail 4", '19-11/16" — save as offcut'], 1, N8); r += 1

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\Bobby Gaming\Desktop\Claude Projects\24x34.5-base-cut-list.xlsx"
wb.save(out)
print("Saved:", out)
